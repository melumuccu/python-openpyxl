import glob
import math
import os
import sys
from datetime import datetime
from functools import reduce

import openpyxl

header_keys = ['date', 'market', 'type', 'price', 'amount', 'total', 'fee', 'fee_coin']
time_regarded_as_same = 15 * 60 # 取りまとめ対象とみなす時間差(秒単位想定)
max_row = 1000 # 最大処理行数

"""
メインロジック
"""
def main():
  # base_url = "/root/file" # コンテナ環境上のパス
  base_url = "/workspaces/python-file-io/sample-files" # devcontainer上のパス
  files = glob.glob(f"{base_url}/*")
  for file_path in files:
    file_name = os.path.basename(file_path)
    wb = openpyxl.load_workbook(f"{base_url}/{file_name}")
    default_sheet = wb['sheet1']

    if 'tmp' in wb.sheetnames:
      wb.remove(wb['tmp'])
    target_sheet = wb.create_sheet(title="tmp")

    initSheet(target_sheet)

    row = 2 # 参照行のindex
    wrire_row = 2 # 対象シートの書き込み対象行のindex
    stocks_row: list = [] # 書き込み保留行
    while True:
      row_i = getBinanceRowData(row, default_sheet)
      row_ii = getBinanceRowData(row+1, default_sheet)
      if row_i == None:
        break
      elif row_ii == None:
        # (最終行想定)
        writeRow(wrire_row, target_sheet, row_i)
        break

      # 行をストック
      stocks_row.append(row_i)

      if not isRegardedAsSame(row_i, row_ii):
        '''============================================
        ここまでストックした行を1行にまとめて書き込みを行う
        ============================================'''
        group: dict[str, list] = {}
        result: dict[str, any] = {}
        # グループの各keyに各行の値をlistとしてまとめる(処理がわかりやすくなるように)
        for key in header_keys:
          group[key] = []
        for stock in stocks_row:
          for key in header_keys:
            group[key].append(stock[key])

        if len(group["date"]) == 1:
          for v in header_keys:
            result[v] = group[v][0]
          writeRow(wrire_row, target_sheet, result)
          stocks_row = []
          wrire_row=wrire_row+1
        else:
          # 1行にまとめる
          result = roundUpIntoRow(group)
          writeRow(wrire_row, target_sheet, result)
          stocks_row = []
          wrire_row=wrire_row+1
      row=row+1
      if row > max_row:
        break
      # END_行単位のループ

    wb.save(f"{base_url}/{file_name}")
    wb.close()
    # END_workbook単位のループ

"""
Excleシートを初期化
・header列を作成
"""
def initSheet(sheet):
  headers = ['Date(UTC)', 'Market', 'Type', 'Price', 'Amount', 'Total', 'Fee', 'Fee Coin']
  for i, v in enumerate(headers):
    sheet.cell(row=1, column=i+1).value = v

"""
Binanace取引所の取引データを1行取得する
"""
def getBinanceRowData(row, sheet):
  date_str = sheet.cell(row, 1).value
  if date_str == None:
    return None
  data = {}
  for i, v in enumerate(header_keys):
    if i == 0:
      data[v] = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
      continue
    data[v] = sheet.cell(row, i+1).value
  return data

"""
１行分記載を行う
"""
def writeRow(row: int, sheet, row_data: dict):
  for i, (k, v) in enumerate(row_data.items()):
    sheet.cell(row, i+1).value = v

"""
以下を全て満たす時、2つの行を1つに取りまとめる判定とする
・取引日時の差が設定時間以下
・通貨ペアが同一
・売買タイプが同一
"""
def isRegardedAsSame(row_1, row_2)->bool:
  delta_seconds = (row_1["date"]-row_2["date"]).total_seconds()
  return delta_seconds < time_regarded_as_same and row_1["market"] == row_2["market"] and row_1["type"] == row_2["type"]

"""
複数行を1行にまとめる
"""
def roundUpIntoRow(data: dict)->dict:
  result: dict[str, any] = {}
  result["date"] = format(data["date"][0], '%Y-%m-%d %H:%M:%S') # 売買日時
  result["market"] = data["market"][0] # 通貨ペア
  result["type"] = data["type"][0] # 売買タイプ
  result["price"] = str( reduce(lambda a, b: float(a) + float(b), data["price"]) / len(data["price"]) ) # 平均価格
  result["amount"] = str( reduce(lambda a, b: float(a) + float(b), data["amount"]) ) # 購入量
  result["total"] = str( reduce(lambda a, b: float(a) + float(b), data["total"]) ) # 合計金額
  result["fee"] = str( reduce(lambda a, b: float(a) + float(b), data["fee"]) ) # 合計手数料
  result["fee_coin"] = data["fee_coin"][0] # 手数料通貨
  return result


"メインロジック起動"
if __name__ == "__main__":
  main()
