import glob
import math
import os
import sys
from datetime import datetime
from functools import reduce

import openpyxl
from openpyxl import Workbook

HEADER_KEYS = ['date', 'market', 'type', 'price', 'amount', 'total', 'fee', 'fee_coin']
TIME_REGARDED_AS_SAME = 15 * 60 # 取りまとめ対象とみなす時間差(秒単位想定)
MAX_ROW = 1000 # 最大処理行数
IS_DEVCONTAINER = os.getcwd() == '/workspaces/python-file-io'
INPUT_URL =  './input' if IS_DEVCONTAINER else '../input' # 入力パス(devcontainerで実行時 <=> コンテナ内で実行時)
OUTPUT_URL = "./output" if IS_DEVCONTAINER else '../output' # 出力パス(devcontainerで実行時 <=> コンテナ内で実行時)

"""
メインロジック
"""
def main():
  files = glob.glob(f"{INPUT_URL}/*")
  for file_path in files:
    file_name = os.path.basename(file_path)
    read_wb = openpyxl.load_workbook(f"{INPUT_URL}/{file_name}", read_only=True)
    read_ws = read_wb['sheet1']
    if os.path.isfile(f"{OUTPUT_URL}/{file_name}"):
      os.remove(f"{OUTPUT_URL}/{file_name}")
    write_wb = Workbook()
    if 'Sheet' in write_wb.sheetnames:
      write_wb.remove(write_wb['Sheet']) # デフォルトのシートを削除
    write_ws = write_wb.create_sheet('sheet1')

    initSheet(write_ws)

    row = 2 # 参照行のindex
    wrire_row = 2 # 対象シートの書き込み対象行のindex
    stocks_row: list = [] # 書き込み保留行
    while True:
      row_i = getBinanceRowData(row, read_ws)
      row_ii = getBinanceRowData(row+1, read_ws)
      if row_i == None:
        break
      elif row_ii == None:
        # (最終行想定)
        writeRow(wrire_row, write_ws, row_i)
        break

      # 行をストック
      stocks_row.append(row_i)

      if not isRegardedAsSame(row_i, row_ii):
        '''============================================
        ここまでストックした行を1行にまとめて書き込みを行う
        ============================================'''
        group: dict[str, list] = {}
        result: dict[str, any] = {}
        # グループの各keyに各行の値をlistとしてまとめる(処理がわかりやすくなるように)
        for key in HEADER_KEYS:
          group[key] = []
        for stock in stocks_row:
          for key in HEADER_KEYS:
            group[key].append(stock[key])

        if len(group["date"]) == 1:
          for v in HEADER_KEYS:
            result[v] = group[v][0]
          writeRow(wrire_row, write_ws, result)
          stocks_row = []
          wrire_row=wrire_row+1
        else:
          # 1行にまとめる
          result = roundUpIntoRow(group)
          writeRow(wrire_row, write_ws, result)
          stocks_row = []
          wrire_row=wrire_row+1
      row=row+1
      if row > MAX_ROW:
        break
      # END_行単位のループ

    # 保存先ディレクトリの存在チェック
    if not os.path.isdir(OUTPUT_URL):
      os.mkdir(OUTPUT_URL, mode=0o777)

    # 空の行の削除
    last_row = write_ws.max_row + 1 #シートの最終行
    for i in reversed(range(1, last_row)):
      if write_ws.cell(row=i, column=1).value == None:
        write_wb.delete_rows(i)

    write_wb.save(f"{OUTPUT_URL}/{file_name}")
    write_wb.close()
    # END_workbook単位のループ

"""
Excleシートを初期化
・header列を作成
"""
def initSheet(sheet):
  headers = ['Date(UTC)', 'Market', 'Type', 'Price', 'Amount', 'Total', 'Fee', 'Fee Coin']
  for i, v in enumerate(headers):
    sheet.cell(row=1, column=i+1).value = v

"""
Binanace取引所の取引データを1行取得する
"""
def getBinanceRowData(row, sheet):
  date_str = sheet.cell(row, 1).value
  if date_str == None:
    return None
  data = {}
  for i, v in enumerate(HEADER_KEYS):
    if i == 0:
      data[v] = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
      continue
    data[v] = sheet.cell(row, i+1).value
  return data

"""
１行分記載を行う
"""
def writeRow(row: int, sheet, row_data: dict):
  for i, (k, v) in enumerate(row_data.items()):
    sheet.cell(row, i+1).value = v

"""
以下を全て満たす時、2つの行を1つに取りまとめる判定とする
・取引日時の差が設定時間以下
・通貨ペアが同一
・売買タイプが同一
"""
def isRegardedAsSame(row_1, row_2)->bool:
  delta_seconds = (row_1["date"]-row_2["date"]).total_seconds()
  return delta_seconds < TIME_REGARDED_AS_SAME and row_1["market"] == row_2["market"] and row_1["type"] == row_2["type"]

"""
複数行を1行にまとめる
"""
def roundUpIntoRow(data: dict)->dict:
  result: dict[str, any] = {}
  result["date"] = format(data["date"][0], '%Y-%m-%d %H:%M:%S') # 売買日時
  result["market"] = data["market"][0] # 通貨ペア
  result["type"] = data["type"][0] # 売買タイプ
  result["price"] = str( reduce(lambda a, b: float(a) + float(b), data["price"]) / len(data["price"]) ) # 平均価格
  result["amount"] = str( reduce(lambda a, b: float(a) + float(b), data["amount"]) ) # 購入量
  result["total"] = str( reduce(lambda a, b: float(a) + float(b), data["total"]) ) # 合計金額
  result["fee"] = str( reduce(lambda a, b: float(a) + float(b), data["fee"]) ) # 合計手数料
  result["fee_coin"] = data["fee_coin"][0] # 手数料通貨
  return result


"メインロジック起動"
if __name__ == "__main__":
  main()
